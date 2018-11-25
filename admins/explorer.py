#!/usr/bin/env Python
# coding=utf-8

from handlers.base import BaseHandler
import json
from admins.decorator import admin_get_auth
from admins.decorator import admin_post_auth
from methods.debug import *
from openpyxl import Workbook
from orm.users_info import UsersInfoModule
from orm.score_info import ScoreInfoModule


# 继承 base.py 中的类 BaseHandler
class AdminExplorerHandler(BaseHandler):
    """
    该类用户文件管理，主要用户导入导出功能
    """
    @admin_get_auth("/admin/explorer", True)
    def get(self):
        user_name = self.get_current_user()
        if user_name is not None:
            self.render("admin/explorer.html", user_name=user_name,
                        controller=self.render_controller,
                        language_mapping=self.language_mapping,
                        )

    @admin_post_auth(False)
    def post(self):
        pass


class FileDownLoadHandler(BaseHandler):
    @admin_get_auth("/admin/explorer", True)
    def get(self):
        table_name = self.get_argument("table_name", "member_table")
        logging.info("table name is %s" % table_name)
        filename = "file_store/"+table_name+".xlsx"
        logging.info("file name is:"+filename)
        if table_name == "member_table":
            self.__write_user_info_to_xlsx(filename)
        elif table_name == "points_table":
            self.__write_score_info_to_xlsx(filename)
        else:

            self.render("admin/file_error.html",
                        controller=self.render_controller,
                        language_mapping=self.language_mapping,
                        )
            return

        self.set_header('Content-Type', 'application/octet-stream')
        self.set_header('Content-Disposition', ('attachment; filename=%s' % filename).encode('utf-8'))

        with open(filename, 'rb') as f:
            while True:
                data = f.read(1024)
                if not data:
                    break

                self.write(data)
        # # 记得有finish哦

        self.finish()

    def __write_user_info_to_xlsx(self, file_name):

        workbook = Workbook()
        booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
        # 存第一行单元格cell(1,1)
        # 1、存标题
        booksheet.cell(1, 1).value = "编号"  # 这个方法索引从1开始
        booksheet.cell(1, 2).value = "用户名"
        booksheet.cell(1, 3).value = "中文名"
        booksheet.cell(1, 4).value = "昵称"
        booksheet.cell(1, 5).value = "密码"
        booksheet.cell(1, 6).value = "邮箱"
        booksheet.cell(1, 7).value = "部门"
        booksheet.cell(1, 8).value = "角色"
        booksheet.cell(1, 9).value = "地址"
        booksheet.cell(1, 10).value = "修改密码次数"

        user_tables = self.db.query(UsersInfoModule).all()

        #  2.存具体信息
        index = 2
        for user in user_tables:
            if user.user_role != "root":
                booksheet.cell(index, 1).value = user.id  # "编号"
                booksheet.cell(index, 2).value = user.user_name  # "用户名"
                booksheet.cell(index, 3).value = user.chinese_name  # "中文名"
                booksheet.cell(index, 4).value = user.nick_name  # "昵称"
                if user.pwd_modified is True:
                    booksheet.cell(index, 5).value = "******"  # "密码"
                else:
                    booksheet.cell(index, 5).value = user.pass_word  # "密码"

                booksheet.cell(index, 6).value = user.email  # "邮箱"
                booksheet.cell(index, 7).value = user.department  # "部门"
                booksheet.cell(index, 8).value = user.user_role  # "角色"
                booksheet.cell(index, 9).value = user.address  # "地址"
                booksheet.cell(index, 10).value = user.change_pwd_count  # "修改密码次数"

                index = index + 1

        workbook.save(file_name)

    def __write_score_info_to_xlsx(self, file_name):

        workbook = Workbook()
        booksheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
        # 存第一行单元格cell(1,1)
        # 1、存标题
        booksheet.cell(1, 1).value = "编号"  # 这个方法索引从1开始
        booksheet.cell(1, 2).value = "用户名"
        booksheet.cell(1, 3).value = "中文名"
        booksheet.cell(1, 4).value = "当前得分"
        booksheet.cell(1, 5).value = "上次得分"

        score_tables = self.db.query(ScoreInfoModule).all()

        #  2.存具体信息
        index = 2
        for score in score_tables:
            booksheet.cell(index, 1).value = score.id  # "编号"
            booksheet.cell(index, 2).value = score.user_name  # "用户名"
            booksheet.cell(index, 3).value = score.chinese_name  # "中文名"
            booksheet.cell(index, 4).value = score.current_scores  # "当前得分"
            booksheet.cell(index, 5).value = score.last_scores  # "上次得分"

            index = index + 1
        workbook.save(file_name)

class FileUpLoadHandler(BaseHandler):
    def post(self):
        ret = {"status": True, "data": "", "error": "succeed"}
        file_metas = self.request.files["file"]               # 获取上传文件信息
        for meta in file_metas:                                 # 循环文件信息
            file_name = meta['filename']                        # 获取文件的名称
            import os                                           # 引入os路径处理模块
            with open(os.path.join('file_store', file_name), 'wb') as up:            # os拼接文件保存路径，以字节码模式打开
                up.write(meta['body'])

        self.write(json.dumps(ret))
